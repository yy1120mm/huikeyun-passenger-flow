#!/usr/bin/env python3
"""
汇客云客流查询 —— 导出 Excel 报表

输入：query_traffic.py 产出的 JSON 记录文件
输出：多 sheet 的 Excel（明细 / 项目汇总 / 日期趋势），指标用中文表头。

用法：
  python query_traffic.py --project P001 --start 2026-07-01 --end 2026-07-07 \
      --format json --output records.json
  python export_excel.py --input records.json --output 客流报表.xlsx --title "7月第一周客流"
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)
from huikeyun_client import load_config  # noqa: E402

DEFAULT_CONFIG = os.path.join(SCRIPT_DIR, "..", "references", "config.yaml")

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def metric_labels(cfg: dict) -> dict:
    reg = cfg.get("metrics", {}) or {}
    return {k: (v.get("label", k) if isinstance(v, dict) else k) for k, v in reg.items()}


def aggregate(records, key_fn, metric_keys):
    acc = defaultdict(lambda: {m: 0.0 for m in metric_keys})
    cnt = defaultdict(int)
    names = {}
    for r in records:
        k = key_fn(r)
        names[k] = r.get("project", "")
        cnt[k] += 1
        for m, v in r.get("metrics", {}).items():
            if v is not None:
                acc[k][m] += v
    return acc, cnt, names


def write_sheet(ws, title, headers, rows, metric_keys):
    ws.append([title])
    ws["A1"].font = TITLE_FONT
    ws.append([])
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for row in rows:
        ws.append(row)
    # 数字格式 & 边框
    for ri in range(4, 4 + len(rows)):
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = BORDER
            if ci > 2:  # 指标列
                cell.number_format = "#,##0.00"
    # 列宽
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(10, len(str(headers[ci - 1])) + 4)
    ws.freeze_panes = "A4"


def main():
    ap = argparse.ArgumentParser(description="汇客云客流 -> Excel")
    ap.add_argument("--input", required=True, help="records.json（来自 query_traffic.py）")
    ap.add_argument("--config", default=DEFAULT_CONFIG)
    ap.add_argument("--output", default="客流报表.xlsx")
    ap.add_argument("--title", default="汇客云客流报表")
    args = ap.parse_args()

    cfg = load_config(args.config)
    labels = metric_labels(cfg)
    with open(args.input, "r", encoding="utf-8") as f:
        records = json.load(f)

    # 只保留 config 里登记过的指标，顺序与 config 一致
    metric_keys = [k for k in (cfg.get("metrics") or {}).keys()
                   if any(k in r.get("metrics", {}) for r in records)] or \
                  list({k for r in records for k in r.get("metrics", {}).keys()})

    wb = Workbook()

    # Sheet 1: 明细
    ws1 = wb.active
    ws1.title = "客流明细"
    headers1 = ["项目", "项目ID", "日期", "时段"] + [labels.get(m, m) for m in metric_keys]
    rows1 = []
    for r in records:
        hour = r.get("hour")
        hour_s = "全天" if hour in (None, "") else f"{int(float(hour)):02d}:00"
        rows1.append([r.get("project", ""), r.get("project_id", ""), r.get("date", ""), hour_s]
                     + [r.get("metrics", {}).get(m) for m in metric_keys])
    write_sheet(ws1, args.title, headers1, rows1, metric_keys)

    # Sheet 2: 项目汇总（求和）
    ws2 = wb.create_sheet("项目汇总")
    acc, cnt, names = aggregate(records, lambda r: r.get("project_id", ""), metric_keys)
    headers2 = ["项目ID", "项目名称", "记录数"] + [labels.get(m, m) for m in metric_keys]
    rows2 = []
    for pid, mvals in acc.items():
        rows2.append([pid, names.get(pid, ""), cnt[pid]] + [mvals[m] for m in metric_keys])
    write_sheet(ws2, args.title + " · 项目汇总", headers2, rows2, metric_keys)

    # Sheet 3: 日期趋势（求和）
    ws3 = wb.create_sheet("日期趋势")
    acc3, cnt3, _ = aggregate(records, lambda r: r.get("date", ""), metric_keys)
    headers3 = ["日期", "记录数"] + [labels.get(m, m) for m in metric_keys]
    rows3 = [[d, cnt3[d]] + [acc3[d][m] for m in metric_keys] for d in sorted(acc3.keys())]
    write_sheet(ws3, args.title + " · 日期趋势", headers3, rows3, metric_keys)

    # 生成时间备注
    ws1.append([])
    ws1.append([f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}"])

    wb.save(args.output)
    print(f"已生成 Excel：{args.output}（明细 {len(records)} 条）")


if __name__ == "__main__":
    main()
